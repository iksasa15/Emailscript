#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
سكربت إرسال الإيميلات من ملف إكسل
يقرأ قائمة الإيميلات من عمود في ملف إكسل ويرسل لها رسالة واحدة
"""

import os
import sys
import smtplib
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

try:
    from openpyxl import load_workbook
except ImportError:
    print("يرجى تثبيت المكتبات: pip install -r requirements.txt")
    sys.exit(1)


# ============ الإعدادات (عدّل حسب احتياجك) ============

# مسار ملف الإكسل
EXCEL_FILE = "emails.xlsx"

# اسم العمود الذي فيه الإيميلات (أو رقم العمود من 1: A=1, B=2, ...)
EMAIL_COLUMN = "email"  # أو استخدم رقم مثل 1

# إعدادات البريد
SMTP_SERVER = "smtp.gmail.com"  # أو smtp.outlook.com لـ Outlook
SMTP_PORT = 587
SENDER_EMAIL = "your_email@gmail.com"

# كلمة مرور التطبيق (لا تضع كلمة مرورك العادية - استخدم "كلمة مرور التطبيق" من إعدادات الحساب)
# الأفضل: ضعها في متغير بيئة EMAIL_PASSWORD
SENDER_PASSWORD = os.environ.get("EMAIL_PASSWORD", "")

# محتوى الرسالة
EMAIL_SUBJECT = "الموضوع هنا"
EMAIL_BODY = """
مرحباً،

هذا نص الرسالة.
يمكنك تعديله كما تريد.

مع التحية
"""

# وقت الانتظار بين كل إيميل (بالثواني) لتجنب حظر الخادم
DELAY_BETWEEN_EMAILS = 2


def load_emails_from_excel(file_path, column):
    """قراءة قائمة الإيميلات من ملف إكسل"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # تحديد عمود الإيميل
    if isinstance(column, int):
        col_index = column
    else:
        # البحث عن اسم العمود في الصف الأول
        headers = [cell.value for cell in ws[1]]
        try:
            col_index = headers.index(column) + 1
        except ValueError:
            raise ValueError(f"العمود '{column}' غير موجود. الأعمدة: {headers}")

    emails = []
    for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
        cell_value = row[0].value
        if cell_value and isinstance(cell_value, str) and "@" in cell_value:
            emails.append(cell_value.strip())

    wb.close()
    return emails


def send_email(to_email, subject, body, sender, password, attachments=None):
    """إرسال إيميل واحد. attachments: قائمة (اسم_الملف, محتوى_bytes)"""
    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = to_email
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "plain", "utf-8"))

    if attachments:
        for filename, content in attachments:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(content)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment", filename=filename)
            msg.attach(part)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(sender, password)
        server.sendmail(sender, to_email, msg.as_string())


def main():
    if not SENDER_PASSWORD:
        print("خطأ: لم يتم تعيين كلمة المرور.")
        print("استخدم: export EMAIL_PASSWORD='كلمة_مرور_التطبيق'")
        print("أو عدّل السكربت وضَع SENDER_PASSWORD (غير مستحسن للأمان)")
        sys.exit(1)

    if not os.path.exists(EXCEL_FILE):
        print(f"خطأ: الملف '{EXCEL_FILE}' غير موجود.")
        print("تأكد من وضع ملف إكسل في نفس مجلد السكربت أو عدّل EXCEL_FILE.")
        sys.exit(1)

    print("جاري قراءة الإيميلات من الإكسل...")
    try:
        emails = load_emails_from_excel(EXCEL_FILE, EMAIL_COLUMN)
    except Exception as e:
        print(f"خطأ في قراءة الملف: {e}")
        sys.exit(1)

    if not emails:
        print("لم يتم العثور على إيميلات صالحة في الملف.")
        sys.exit(1)

    print(f"تم العثور على {len(emails)} إيميل. جاري الإرسال...\n")

    success = 0
    failed = 0

    for i, email in enumerate(emails, 1):
        try:
            send_email(
                email,
                EMAIL_SUBJECT,
                EMAIL_BODY,
                SENDER_EMAIL,
                SENDER_PASSWORD,
            )
            success += 1
            print(f"  [{i}/{len(emails)}] تم الإرسال إلى: {email}")
        except Exception as e:
            failed += 1
            print(f"  [{i}/{len(emails)}] فشل {email}: {e}")

        if i < len(emails) and DELAY_BETWEEN_EMAILS > 0:
            time.sleep(DELAY_BETWEEN_EMAILS)

    print(f"\nانتهى. نجح: {success} | فشل: {failed}")


if __name__ == "__main__":
    main()
