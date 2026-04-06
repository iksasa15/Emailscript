#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""خادم ويب لواجهة إرسال الإيميلات"""

import json
import os
import tempfile
import time
from flask import Flask, request, jsonify, send_from_directory, send_file, Response, stream_with_context

from io import BytesIO
from send_emails import load_emails_from_excel, send_email, SMTP_SERVER, SMTP_PORT

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

app = Flask(__name__, static_folder="static", static_url_path="")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB


def sse_message(data):
    return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"


@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/preview", methods=["POST"])
def preview():
    """معاينة قائمة الإيميلات من ملف الإكسل دون إرسال"""
    file = request.files.get("emails_file")
    if not file or file.filename == "":
        return jsonify({"ok": False, "error": "يرجى رفع ملف الإيميلات (إكسل)"}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"ok": False, "error": "الملف يجب أن يكون إكسل (.xlsx)"}), 400
    email_column = request.form.get("email_column", "email").strip() or "email"
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        try:
            emails = load_emails_from_excel(tmp_path, email_column)
        finally:
            os.unlink(tmp_path)
        return jsonify({"ok": True, "emails": emails, "count": len(emails)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/export-results", methods=["POST"])
def export_results():
    """تصدير نتائج الإرسال كملف إكسل"""
    data = request.get_json()
    if not data or "results" not in data:
        return jsonify({"error": "بيانات غير صالحة"}), 400
    results = data["results"]
    if not Workbook:
        return jsonify({"error": "openpyxl غير متوفر"}), 500
    wb = Workbook()
    ws = wb.active
    ws.title = "نتائج الإرسال"
    ws["A1"] = "الإيميل"
    ws["B1"] = "الحالة"
    ws["C1"] = "رسالة الخطأ"
    for i, row in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=row.get("email", ""))
        ws.cell(row=i, column=2, value=row.get("status", ""))
        ws.cell(row=i, column=3, value=row.get("message", ""))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="نتائج_الإرسال.xlsx",
    )


@app.route("/send", methods=["POST"])
def send():
    sender_email = request.form.get("email", "").strip()
    sender_password = request.form.get("password", "")
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    email_column = request.form.get("email_column", "email").strip() or "email"
    delay = float(request.form.get("delay", "1.5") or "1.5")

    if not sender_email or not sender_password:
        return jsonify({"ok": False, "error": "الإيميل وكلمة المرور مطلوبان"}), 400
    if not subject:
        return jsonify({"ok": False, "error": "عنوان الرسالة مطلوب"}), 400

    file = request.files.get("emails_file")
    if not file or file.filename == "":
        return jsonify({"ok": False, "error": "يرجى رفع ملف الإيميلات (إكسل)"}), 400
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"ok": False, "error": "الملف يجب أن يكون إكسل (.xlsx)"}), 400

    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        try:
            emails = load_emails_from_excel(tmp_path, email_column)
        finally:
            os.unlink(tmp_path)

        if not emails:
            return jsonify({"ok": False, "error": "لم يتم العثور على إيميلات صالحة في الملف"}), 400

        attachments = []
        for f in request.files.getlist("attachments"):
            if f and f.filename:
                attachments.append((f.filename, f.read()))
        attachments = attachments if attachments else None
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    def generate():
        success = 0
        failed = 0
        total = len(emails)
        for i, to_email in enumerate(emails):
            yield sse_message({
                "status": "sending",
                "email": to_email,
                "index": i + 1,
                "total": total,
            })
            try:
                send_email(
                    to_email,
                    subject,
                    body,
                    sender_email,
                    sender_password,
                    attachments=attachments,
                )
                success += 1
                yield sse_message({
                    "status": "sent",
                    "email": to_email,
                    "index": i + 1,
                    "total": total,
                })
            except Exception as e:
                failed += 1
                yield sse_message({
                    "status": "error",
                    "email": to_email,
                    "message": str(e),
                    "index": i + 1,
                    "total": total,
                })
            if i < total - 1 and delay > 0:
                time.sleep(delay)
        yield sse_message({
            "status": "done",
            "success": success,
            "failed": failed,
            "total": total,
        })

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    default_host = "0.0.0.0" if "PORT" in os.environ else "127.0.0.1"
    host = os.environ.get("FLASK_HOST", default_host)
    debug = os.environ.get("FLASK_DEBUG", "").lower() in ("1", "true", "yes")
    app.run(host=host, port=port, debug=debug)
