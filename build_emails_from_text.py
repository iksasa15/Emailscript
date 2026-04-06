#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""استخراج الإيميلات من نص خام وحفظها في emails.xlsx"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

# تعبير نمطي للإيميلات الشائعة
EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
    re.UNICODE,
)


def normalize_email(m: str) -> str:
    m = m.strip().rstrip(".,;:!?")
    # إزالة نقطة زائدة في نهاية النطاق مثل .com.
    while m.endswith(".") and m.count("@") == 1:
        m = m[:-1]
    return m.strip()


def extract_emails(text: str) -> list[str]:
    seen = set()
    out = []
    for raw in EMAIL_RE.findall(text):
        e = normalize_email(raw)
        if "@" not in e:
            continue
        el = e.lower()
        if el not in seen:
            seen.add(el)
            out.append(e)
    return out


def emails_from_list_file(path: Path) -> list[str]:
    """سطر واحد = إيميل (بدون أسماء). تكرار يُزال مع الحفاظ على الترتيب."""
    seen: set[str] = set()
    out: list[str] = []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = line.strip()
        if "@" not in line or line.startswith("#"):
            continue
        key = line.lower()
        if key not in seen:
            seen.add(key)
            out.append(line)
    return out


def main():
    base = Path(__file__).parent
    list_path = base / "emails_list.txt"
    if list_path.exists():
        emails = emails_from_list_file(list_path)
    else:
        parts = sorted(base.glob("emails_raw*.txt"))
        if parts:
            text = "".join(p.read_text(encoding="utf-8", errors="replace") for p in parts)
        else:
            raw_path = base / "emails_raw.txt"
            if not raw_path.exists():
                print(f"ضع إما emails_list.txt أو النص في {raw_path} أو emails_raw1.txt ...")
                sys.exit(1)
            text = raw_path.read_text(encoding="utf-8", errors="replace")
        emails = extract_emails(text)

    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"
    ws["A1"] = "email"
    for i, em in enumerate(emails, start=2):
        ws.cell(row=i, column=1, value=em)

    out_xlsx = Path(__file__).parent / "emails.xlsx"
    wb.save(out_xlsx)
    print(f"تم حفظ {len(emails)} إيميل فريد في {out_xlsx}")


if __name__ == "__main__":
    main()
